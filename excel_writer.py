import argparse
import pathlib

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class _WorkbookEditor:
    def __init__(self, args: argparse.Namespace):
        self._args = args
        self._workbook: Workbook = openpyxl.Workbook()

        self._clinical_sheetname = "Clinically Relevant Proteins"
        self._all_proteins_sheetname = "All Proteins"
        self._headings = ["SDC", "SDC-C18", "Urea", "Urea-C18"]
        self._subheadings = [
            "Dried\nAverage",
            "Dried\n%CV",
            "Liquid\nAverage",
            "Liquid\n%CV",
            "D/L\nRatio",
        ]

        self._first_setup: bool = self._setup_workbook()

        if self.first_setup:
            self._workbook[self._all_proteins_sheetname].delete_cols(3)
            self._remerge_cells()

    def _setup_workbook(self) -> bool:
        if pathlib.Path.exists(pathlib.Path(self._args.excel)):
            self._workbook = openpyxl.load_workbook(self._args.excel)
            return False
        else:
            self._workbook = openpyxl.Workbook()

            del self._workbook["Sheet"]

            self._workbook.create_sheet(self.clinical_sheetname)
            self._workbook.create_sheet(self.all_proteins_sheetname)

            self._set_formatting()
            self._write_headings()

            return True

    def _set_formatting(self) -> None:
        """
        This function is responsible for any formatting required of the excel sheet
        This includes setting alignments, freezing rows for headers, and setting borders
        :return:
        """
        # Freeze top two rows for headers
        self._workbook[self.clinical_sheetname].freeze_panes = "A3"
        self._workbook[self._all_proteins_sheetname].freeze_panes = "A3"

        # Set alignment
        # Iterate through each worksheet
        for sheet in self._workbook.worksheets:
            for row in range(1, 3):
                for col in range(1, 24):
                    sheet.cell(row, col).alignment = Alignment(
                        wrapText=True, horizontal="center", vertical="center"
                    )

        medium_left_border = Border(left=Side(style="medium"))
        medium_bottom_border = Border(bottom=Side(style="medium"))

        # Add vertical borders between each experiment (SDC, SDC-C18, etc.)
        for row in range(3, 500):
            for column in range(4, 25, 5):
                # Set left borders for clinical worksheet
                self._workbook[self._clinical_sheetname].cell(
                    row=row, column=column
                ).border = medium_left_border

                # Set borders for all proteins worksheet
                self._workbook[self._all_proteins_sheetname].cell(
                    row=row, column=column
                ).border = medium_left_border

        # Add horizontal border below subheading
        for column in range(1, 24):
            self._workbook[self._clinical_sheetname].cell(
                row=2, column=column
            ).border = medium_bottom_border

            self._workbook[self._all_proteins_sheetname].cell(
                row=2, column=column
            ).border = medium_bottom_border

        # Set width for clinical worksheet column A and B
        # From: https://stackoverflow.com/a/35790441
        self._workbook[self._clinical_sheetname].column_dimensions["A"].width = 51
        self._workbook[self._clinical_sheetname].column_dimensions["B"].width = 12

        # Set width for all proteins worksheet column A and B
        # Do not set column A to "max" value, as this could be very long
        self._workbook[self._all_proteins_sheetname].column_dimensions["A"].width = 70
        self._workbook[self._all_proteins_sheetname].column_dimensions["B"].width = 35

    def _write_headings(self) -> None:
        """
        This function will write the heading values for each worksheet
        It attempts to do this in the least amount of copy-paste code as possible
        :return:
        """
        # Get worksheets to write data that is independent between the two
        clinical_sheet: Worksheet = self._workbook[self.clinical_sheetname]
        all_proteins_sheet: Worksheet = self._workbook[self.all_proteins_sheetname]

        # Set clinical worksheet headings
        clinical_sheet["A1"] = "Clinically Relevant"
        clinical_sheet["C1"] = "Typical\nPlasma\nConc"
        clinical_sheet.merge_cells("C1:C2")

        # Set All Proteins sheet headings
        all_proteins_sheet["A1"] = "Identified Proteins"

        # Iterate through each sheet in the workbook
        for sheet in self._workbook.worksheets:

            # Write values that are consistent between the two sheets
            sheet["A2"] = "Protein\nName"
            sheet["B2"] = "Majority\nProtein\nID"
            sheet.merge_cells("A1:B1")

            # Start writing the top-most headings
            for i, col_num in enumerate(range(4, 24, 5)):

                sheet.cell(row=1, column=col_num, value=self._headings[i])
                sheet.merge_cells(
                    start_row=1, end_row=1, start_column=col_num, end_column=col_num + 4
                )

                # Write subheadings (dried average, dried %CV, liquid average, etc.)
                for sub_col_num in range(0, 5):
                    sheet.cell(
                        row=2,
                        column=col_num + sub_col_num,
                        value=self._subheadings[sub_col_num],
                    )

    def _remerge_cells(self):
        """
        This function will re-merge cells after deleting a column

        From: https://stackoverflow.com/questions/58412906

        :return: None
        """
        delete_index = 3
        for merged_cells in self._workbook[self.all_proteins_sheetname].merged_cells:
            if delete_index < merged_cells.min_col:
                merged_cells.shift(col_shift=-1)
            elif delete_index <= merged_cells.max_col:
                merged_cells.shrink(right=1)

    def get_column_write_start(self, sheet_title: str) -> int:
        """
        This function is responsible for determining which column should be written to when adding data to the excel file
        :param sheet_title:
        :return:
        """

        # Determine if we are running Direct or C18 data
        if str(self._args.method).lower() == "direct":
            start_col = 4
        else:
            start_col = 9

        # If this is a urea experiment, we need to go up by 10 columns
        if str(self._args.experiment).lower() == "urea":
            start_col += 10

        # If we are NOT doing first setup, and we ARE on the All Proteins sheet, decrease by one
        # This is because the third column of ALl Proteins is deleted
        if not self.first_setup and sheet_title == self.all_proteins_sheetname:
            start_col -= 1

        return start_col

    @property
    def workbook(self) -> Workbook:
        return self._workbook

    @property
    def clinical_sheetname(self) -> str:
        return self._clinical_sheetname

    @property
    def all_proteins_sheetname(self) -> str:
        return self._all_proteins_sheetname

    @property
    def first_setup(self) -> bool:
        return self._first_setup

    @property
    def headings(self) -> list[str]:
        return self._headings

    @property
    def excel_subheadings(self) -> list[str]:
        return self._subheadings


class ClinicallyRelevant:
    def __init__(self, data_frame: pd.DataFrame, args: argparse.Namespace):
        self._args = args
        self._editor = _WorkbookEditor(args)
        self._workbook: Workbook = self._editor.workbook
        self._sheet: Worksheet = self._workbook[self._editor.clinical_sheetname]
        self._dataframe: pd.DataFrame = data_frame[data_frame["relevant"]].reset_index(
            drop=True
        )

        if self._editor.first_setup:
            self._write_clinical_name_id()

        self._write_clinical_data()
        self._workbook.save(self._args.excel)

    def _write_clinical_name_id(self):
        """
        This function will write ONLY the clinicaly name and ID contained in the clinically_relevant.tsv file to columns 1 and 2 in the excel file
        :return:
        """
        clinical_df: pd.DataFrame = pd.read_csv("clinically_relevant.tsv", sep="\t")
        clinical_df.sort_values(
            "protein_name",
            ignore_index=True,
            inplace=True,
            # Sort by lowercase. From: https://stackoverflow.com/a/63141564
            key=lambda col: col.str.lower(),
        )

        for i, (name, protein_id, concentration) in enumerate(
            zip(
                clinical_df["protein_name"],
                clinical_df["protein_id"],
                clinical_df["expected_concentration [log10(pg/ml)]"],
            )
        ):
            self._sheet.cell(row=i + 3, column=1, value=name)
            self._sheet.cell(row=i + 3, column=2, value=protein_id)

            if int(concentration) != -1:
                self._sheet.cell(row=i + 3, column=3, value=concentration)

    def _write_clinical_data(self):
        """
        This function will match clinically relevant MaxQuant data lines located in the Clinically Relevant Proteins file
        :return: None
        """

        start_col = self._editor.get_column_write_start(self._sheet.title)

        for i, (
            clinical_id,
            dried_average,
            dried_variation,
            liquid_average,
            liquid_variation,
            ratio,
        ) in enumerate(
            zip(
                self._dataframe["clinical_id"],
                self._dataframe["dried_average"],
                self._dataframe["dried_variation"],
                self._dataframe["liquid_average"],
                self._dataframe["liquid_variation"],
                self._dataframe["dried_liquid_ratio"],
            )
        ):
            for j, row_contents in enumerate(
                self._sheet.iter_rows(min_row=3, min_col=2, max_col=2)
            ):
                row_index = j + 3
                excel_id: str = row_contents[0].value

                if clinical_id == excel_id:
                    self._sheet.cell(
                        row=row_index, column=start_col, value=dried_average
                    )

                    self._sheet.cell(
                        row=row_index, column=start_col + 1, value=dried_variation
                    )
                    self._sheet.cell(
                        row=row_index, column=start_col + 2, value=liquid_average
                    )
                    self._sheet.cell(
                        row=row_index, column=start_col + 3, value=liquid_variation
                    )

                    self._sheet.cell(row=row_index, column=start_col + 4, value=ratio)

                    break


class AllProteins:
    def __init__(self, data_frame: pd.DataFrame, args: argparse.Namespace):
        self._args = args
        self._editor = _WorkbookEditor(args)
        self._workbook: Workbook = self._editor.workbook

        self._sheet: Worksheet = self._workbook[self._editor.all_proteins_sheetname]

        self._dataframe: pd.DataFrame = data_frame[
            data_frame["relevant"] == False
        ].reset_index(drop=True)

        # Generate dataframes
        self._ingested_df: pd.DataFrame = self._ingest_protein_data()
        self._incoming_frame: pd.DataFrame = self._format_incoming_frame()
        self._merged_frame: pd.DataFrame = pd.concat(
            [self._ingested_df, self._incoming_frame]
        )

        # Sort using protein name, then index
        # From: https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.sort_values.html
        self._sorted_df: pd.DataFrame = self._merged_frame.sort_values(
            ["protein_name", "index"]
        ).reset_index(drop=True)

        self._write_data()
        self._workbook.save(self._args.excel)

    def _ingest_protein_data(self) -> pd.DataFrame:
        """
        This function will be responsible for returning a dataframe containing the information from the current excel file
        :return:
        """

        ingested_data: dict[str, list] = {
            "protein_name": [],
            "protein_id": [],
            "method": [],
            "experiment": [],
            "value": [],
            "index": [],
        }

        # Get protein data
        for i, row in enumerate(
            self._sheet.iter_rows(min_row=3, min_col=1, max_col=22)
        ):

            # Indexes are required to be able to ensure the data frame can be sorted
            # This is used in case there are two "protein_name" values that clash
            # Index will be used to ensure dried/liquid averages, etc. will map to the correct location
            ingested_data["index"].extend([0, 1, 2, 3, 4] * 4)

            for j, cell in enumerate(row):
                # We don't want to actually iterate through the protein names & IDs, we just want the information contained in the cells

                if j >= 2:
                    ingested_data["protein_name"].append(row[0].value)
                    ingested_data["protein_id"].append(row[1].value)
                    try:
                        ingested_data["value"].append(float(cell.value))
                    except TypeError:
                        ingested_data["value"].append(0.0)

                # Direct SDC values
                if 2 <= j <= 6:
                    ingested_data["method"].append("Direct")
                    ingested_data["experiment"].append("SDC")

                # C18 SDC values
                elif 7 <= j <= 11:
                    ingested_data["method"].append("C18")
                    ingested_data["experiment"].append("SDC")

                # Direct Urea values
                elif 12 <= j <= 16:
                    ingested_data["method"].append("Direct")
                    ingested_data["experiment"].append("Urea")

                # C18 Urea values
                elif 17 <= j <= 21:
                    ingested_data["method"].append("C18")
                    ingested_data["experiment"].append("Urea")

        ingested_df: pd.DataFrame = pd.DataFrame.from_dict(ingested_data)

        # Create a multi-index from the dataframe
        # From: https://pandas.pydata.org/pandas-docs/stable/user_guide/advanced.html
        # From:https://towardsdatascience.com/how-to-use-multiindex-in-pandas-to-level-up-your-analysis-aeac7f451fce
        # ingested_df.set_index(["protein_name, "protein_id", "method", "experiment"], inplace=True)

        return ingested_df

    def _format_incoming_frame(self) -> pd.DataFrame:
        """
        This function is responsible for creating a multi-index dataframe from the incoming data frame
        This will make it easier to merge the values from excel and the new, incoming information
        :return:
        """
        required_data: pd.DataFrame = self._dataframe[
            (self._dataframe["relevant"] == False)
            & (self._dataframe["protein_name"] != "")
        ]
        required_data = required_data[
            [
                "protein_name",
                "protein_id",
                "dried_average",
                "dried_variation",
                "liquid_average",
                "liquid_variation",
                "dried_liquid_ratio",
            ]
        ]
        data_matches: dict[str, list] = {
            "protein_name": [],
            "protein_id": [],
            "method": [],
            "experiment": [],
            "value": [],
            "index": [],
        }

        for i, (
            name,
            protein_id,
            dried_average,
            dried_variation,
            liquid_average,
            liquid_variation,
            ratio,
        ) in enumerate(
            zip(
                required_data["protein_name"],
                required_data["protein_id"],
                required_data["dried_average"],
                required_data["dried_variation"],
                required_data["liquid_average"],
                required_data["liquid_variation"],
                required_data["dried_liquid_ratio"],
            )
        ):

            data_matches["protein_name"].extend([name] * 5)
            data_matches["protein_id"].extend([protein_id] * 5)
            data_matches["method"].extend([self._args.method] * 5)
            data_matches["experiment"].extend([self._args.experiment] * 5)
            data_matches["value"].extend(
                [
                    int(dried_average),
                    float(dried_variation),
                    int(liquid_average),
                    float(liquid_variation),
                    float(ratio),
                ]
            )

            # Index values are required to ensure the "value" column can be sorted properly
            data_matches["index"].extend([0, 1, 2, 3, 4])

        incoming_df: pd.DataFrame = pd.DataFrame.from_dict(data_matches)
        # incoming_df.set_index(["protein_name, "protein_id", "method", "experiment"], inplace=True)
        return incoming_df

    def _write_data(self):
        """
        This function is responsible for taking the merged data frame and writing data to the appropriate location
        :return:
        """

        row_index = 3
        for i, (
            protein_name,
            protein_id,
            method,
            experiment,
            value,
            index,
        ) in enumerate(
            zip(
                self._sorted_df["protein_name"],
                self._sorted_df["protein_id"],
                self._sorted_df["method"],
                self._sorted_df["experiment"],
                self._sorted_df["value"],
                self._sorted_df["index"],
            )
        ):

            # Skip unknown protein names
            if protein_name is None:
                continue

            # Only get first protein ID if protein_id is not none
            if protein_id is not None:
                max_protein_id = protein_id.split(";")[0]
                additional_ids = protein_id.split(";")[1:]
            else:
                max_protein_id = protein_id
                additional_ids = ""

            # Default to SDC column for direct and c18
            if method.lower() == "direct":
                column_index = 3
            else:
                column_index = 8

            # Only need to modify initial column index if we are writing to urea locations
            if experiment.lower() == "urea":
                column_index += 10

            # Modify dried/liquid average, etc based on index value
            column_index += index

            # Don't want to write 0.00 values, write empty string instead
            if value == 0:
                value = ""

            self._sheet.cell(row=row_index, column=1, value=protein_name)
            self._sheet.cell(row=row_index, column=2, value=protein_id)
            self._sheet.cell(row=row_index, column=column_index, value=value)

            # Write to next row if current protein name differs from next protein name
            try:
                next_name = self._sorted_df.loc[i + 1, "protein_name"]
                if protein_name != next_name:
                    row_index += 1

            # next_name will error-out on the last value, must account for this
            except KeyError:
                pass
